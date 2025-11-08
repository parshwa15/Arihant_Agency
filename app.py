from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
from io import BytesIO
import uuid
import calendar
from datetime import datetime
from typing import Dict, Any, List, Optional

from openpyxl import load_workbook
import xlsxwriter

from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
)
from reportlab.lib.styles import getSampleStyleSheet

app = Flask(__name__, static_url_path="/static", template_folder="templates")
CORS(app)

# In-memory per upload
UPLOADS: Dict[str, Dict[str, Any]] = {}
# {
#   upload_id: {
#       "headers": [..],
#       "rows": [ {header:value(str for display)} ],
#       "rows_raw": [ {header:value(raw)} ],
#       "dealer_name_col": str|None,
#       "dealer_code_col": str|None,
#       "month_col": str|None
#   }
# }

# ---------- Helpers (no pandas) ----------

def _normalize_headers(headers: List[str]) -> List[str]:
    return [("" if h is None else str(h).strip()) for h in headers]

def _read_excel(file_storage) -> Dict[str, Any]:
    wb = load_workbook(file_storage, data_only=True)
    ws = wb.active  # first sheet
    headers = _normalize_headers([c.value for c in ws[1]])
    rows_raw = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rec = {}
        for i, h in enumerate(headers):
            rec[h] = r[i] if i < len(r) else None
        rows_raw.append(rec)
    return {"headers": headers, "rows_raw": rows_raw}

def _detect_dealer_name_col(headers: List[str]) -> Optional[str]:
    low = {h.lower(): h for h in headers}
    # exact “dealer name”
    for k, v in low.items():
        cleaned = k.replace("_"," ").replace("-"," ").strip()
        if cleaned == "dealer name":
            return v
    # contains both words
    for k, v in low.items():
        if "dealer" in k and "name" in k:
            return v
    # party name fallback
    for k, v in low.items():
        if "party" in k and "name" in k:
            return v
    return None

def _detect_dealer_code_col(headers: List[str]) -> Optional[str]:
    low = {h.lower(): h for h in headers}
    for k, v in low.items():
        cleaned = k.replace("_"," ").replace("-"," ").strip()
        if cleaned in ("dealer code", "dealercode", "code"):
            return v
    for k, v in low.items():
        if "dealer" in k and "code" in k:
            return v
    return None

def _detect_month_col(headers: List[str]) -> Optional[str]:
    """Prefer explicit MONTH column; else any date-like column such as SALE_DATE."""
    low = {h.lower(): h for h in headers}
    for candidate in ["month", "mnth", "billing month", "bill month"]:
        if candidate in low:
            return low[candidate]
    for h in headers:
        cl = h.lower()
        if any(x in cl for x in ["sale_date", "sale date", "date", "invoice", "bill"]):
            return h
    return None

# ---- Date parsing tuned for 20250812 (YYYYMMDD) ----
def _try_parse_date(val):
    # already a datetime
    if isinstance(val, datetime):
        return val
    # integers like 20250812
    if isinstance(val, int):
        s = str(val)
        if len(s) == 8:
            try:
                return datetime.strptime(s, "%Y%m%d")
            except Exception:
                return val
        return val
    # floats like 20250812.0
    if isinstance(val, float):
        s = str(int(val))
        if len(s) == 8:
            try:
                return datetime.strptime(s, "%Y%m%d")
            except Exception:
                return val
        return val
    # strings
    if isinstance(val, str):
        s = val.strip()
        if s.isdigit() and len(s) == 8:
            try:
                return datetime.strptime(s, "%Y%m%d")
            except Exception:
                pass
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d.%m.%Y"):
            try:
                return datetime.strptime(s, fmt)
            except Exception:
                pass
    return val

def _month_name_from_value(val) -> Optional[str]:
    dt = _try_parse_date(val)
    if isinstance(dt, datetime):
        return dt.strftime("%B")
    if isinstance(val, str):
        name = val.strip()
        cap = name.capitalize()
        if cap in list(calendar.month_name)[1:]:
            return cap
        if cap[:3] in list(calendar.month_abbr)[1:]:
            idx = list(calendar.month_abbr).index(cap[:3])
            return calendar.month_name[idx]
    return None

def _to_display(val) -> str:
    dt = _try_parse_date(val)
    if isinstance(dt, datetime):
        return dt.strftime("%d/%m/%Y")
    return "" if val is None else str(val)

def _draw_footer(canvas, doc):
    """Footer watermark on every page."""
    canvas.saveState()
    canvas.setFont("Helvetica", 9)
    canvas.setFillGray(0.55)  # light gray
    footer_text = "By Arihant Agency, Sangamner"
    page_width, _ = landscape(A4)
    canvas.drawCentredString(page_width / 2, 12, footer_text)
    canvas.restoreState()

# ---------- Routes ----------

@app.get("/")
def home():
    return render_template("index.html")

@app.post("/upload")
def upload():
    f = request.files.get("file")
    if not f:
        return jsonify(success=False, error="No file uploaded"), 400

    try:
        parsed = _read_excel(f)
    except Exception as e:
        return jsonify(success=False, error=f"Failed to read Excel: {e}"), 400

    headers = parsed["headers"]
    rows_raw = parsed["rows_raw"]

    # display-friendly copy
    rows = []
    for rec in rows_raw:
        rows.append({h: _to_display(rec.get(h)) for h in headers})

    dealer_name_col = _detect_dealer_name_col(headers)
    dealer_code_col = _detect_dealer_code_col(headers)
    month_col = _detect_month_col(headers)

    upload_id = uuid.uuid4().hex
    UPLOADS[upload_id] = {
        "headers": headers,
        "rows": rows,
        "rows_raw": rows_raw,
        "dealer_name_col": dealer_name_col,
        "dealer_code_col": dealer_code_col,
        "month_col": month_col
    }

    # Dealers dropdown
    dealers: List[str] = []
    if dealer_name_col:
        dealers = sorted({ (rec.get(dealer_name_col) or "").strip() for rec in rows })

    # Months dropdown (Jan→Dec order) from month_col
    months_set = set()
    if month_col:
        for r in rows_raw:
            m = _month_name_from_value(r.get(month_col))
            if m: months_set.add(m)
    months = [m for m in list(calendar.month_name)[1:] if m in months_set]  # ordered

    return jsonify(
        success=True,
        message="Sheet loaded successfully ✅",
        upload_id=upload_id,
        dealer_name_col=dealer_name_col,
        dealer_code_col=dealer_code_col,
        month_col=month_col,
        dealers=dealers,
        months=months
    )

def _filter(upload_id: str, dealer_value: str, month_value: str):
    if upload_id not in UPLOADS:
        return None, None, "Invalid upload_id"
    data = UPLOADS[upload_id]
    headers = data["headers"]
    rows = data["rows"]
    rows_raw = data["rows_raw"]
    dealer_name_col = data["dealer_name_col"]
    dealer_code_col = data["dealer_code_col"]
    month_col = data["month_col"]

    # Filter by dealer (optional)
    if dealer_value:
        def match_dealer(rec): return (rec.get(dealer_name_col) or "").strip() == dealer_value
        rows     = [r for r in rows if match_dealer(r)]
        rows_raw = [r for r in rows_raw if match_dealer(r)]

    # Dealer code for header (from first matching row)
    dealer_code_value = None
    if dealer_code_col and rows_raw:
        dc = rows_raw[0].get(dealer_code_col)
        dealer_code_value = "" if dc is None else str(dc)

    # Filter by month (optional)
    eff_month = "All"
    if month_value and month_value != "ALL" and month_col:
        def keep(r): return _month_name_from_value(r.get(month_col)) == month_value
        filtered_raw = [r for r in rows_raw if keep(r)]
        rows = [{h: _to_display(rr.get(h)) for h in headers} for rr in filtered_raw]
        eff_month = month_value

    return rows, {"dealer_code": dealer_code_value, "month": eff_month}, None

@app.get("/dealer-data")
def dealer_data():
    upload_id = request.args.get("upload_id")
    dealer_value = request.args.get("dealer","")
    month_value = request.args.get("month","ALL")
    rows, meta, err = _filter(upload_id, dealer_value, month_value)
    if err:
        return jsonify(success=False, error=err), 404
    return jsonify(success=True, rows=rows, total=len(rows), dealer_code=meta.get("dealer_code"), month_label=meta.get("month","All"))

@app.get("/export/excel")
def export_excel():
    upload_id = request.args.get("upload_id")
    dealer_value = request.args.get("dealer","")
    month_value = request.args.get("month","ALL")
    rows, _, err = _filter(upload_id, dealer_value, month_value)
    if err:
        return jsonify(success=False, error=err), 404

    output = BytesIO()
    headers = UPLOADS[upload_id]["headers"]
    wb = xlsxwriter.Workbook(output, {'in_memory': True})
    ws = wb.add_worksheet("DealerData")
    # headers
    for c, h in enumerate(headers):
        ws.write(0, c, h)
    # rows
    for r_idx, row in enumerate(rows, start=1):
        for c, h in enumerate(headers):
            ws.write(r_idx, c, row.get(h, ""))
    wb.close()
    output.seek(0)
    filename = f"dealer_data_{(dealer_value or 'all').replace(' ','_')}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/export/pdf")
def export_pdf():
    upload_id = request.args.get("upload_id")
    dealer_value = request.args.get("dealer","")
    month_value = request.args.get("month","ALL")
    rows, meta, err = _filter(upload_id, dealer_value, month_value)
    if err:
        return jsonify(success=False, error=err), 404

    # Build table data, dropping dealer name/code from the body table
    headers_all = [h for h in UPLOADS[upload_id]["headers"]]
    dealer_name_col = UPLOADS[upload_id]["dealer_name_col"]
    dealer_code_col = UPLOADS[upload_id]["dealer_code_col"]
    headers = [h for h in headers_all if h not in (dealer_name_col, dealer_code_col)]

    # Use Paragraph to allow wrapping
    styles = getSampleStyleSheet()
    cell_style = styles["BodyText"]
    cell_style.fontSize = 8
    cell_style.leading = 10
    head_style = styles["BodyText"]
    head_style.fontSize = 9
    head_style.leading = 11

    head_row = [Paragraph(str(h), head_style) for h in headers]
    body_rows = []
    for r in rows:
        body_rows.append([Paragraph("" if r.get(h) is None else str(r.get(h)), cell_style) for h in headers])

    data = [head_row] + body_rows

    # Document & width math
    page_w, page_h = landscape(A4)
    left, right, top, bottom = 18, 18, 18, 18
    usable_w = page_w - left - right

    # Fill the page width evenly
    num_cols = max(1, len(headers))
    col_width = usable_w / num_cols
    col_widths = [col_width] * num_cols

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=left, rightMargin=right, topMargin=top, bottomMargin=bottom
    )

    elements = []
    title = Paragraph(f"Dealer Report - {dealer_value or 'All Dealers'}", styles["Heading2"])
    meta_line = Paragraph(
        f"<b>Dealer Code:</b> {meta.get('dealer_code') or '-'} &nbsp;&nbsp; "
        f"<b>Month:</b> {meta.get('month','All')}",
        styles["Normal"]
    )
    elements.append(title)
    elements.append(meta_line)
    elements.append(Spacer(1, 8))

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#dddd12")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('FONTSIZE', (0,1), (-1,-1), 8),
        ('LEADING', (0,1), (-1,-1), 10),
        ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor("#f1f5f9")]),
    ]))
    elements.append(table)

    # Footer on every page
    doc.build(elements, onFirstPage=_draw_footer, onLaterPages=_draw_footer)

    buffer.seek(0)
    filename = f"dealer_data_{(dealer_value or 'all').replace(' ','_')}.pdf"
    return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
